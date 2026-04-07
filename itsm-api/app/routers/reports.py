"""Reports router: real-time stats, trends, ratings, and CSV export."""
import csv
import io
import logging
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, time as dt_time, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from ..database import get_db
from .. import gitlab_client
from ..models import DailyStatsSnapshot, Rating, SLARecord, TimeEntry
from ..rbac import require_agent
from ..redis_client import get_redis

_CACHE_TTL_STATS = 60        # current-stats: 1분
_CACHE_TTL_BREAKDOWN = 60    # breakdown: 1분
_CACHE_TTL_SLA_DASH = 120    # sla-dashboard: 2분
_CACHE_TTL_AGENT_PERF = 120  # agent-performance: 2분 (GitLab 페이지네이션 비용 큼)

router = APIRouter(prefix="/reports", tags=["reports"])
logger = logging.getLogger(__name__)


def _sanitize_csv_cell(value: str) -> str:
    """Prevent CSV formula injection (Excel/LibreOffice)."""
    if value and value[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + value  # prefix with single quote
    return value


@router.get("/current-stats")
def get_current_stats(
    from_date: Optional[date] = Query(default=None, alias="from"),
    to_date: Optional[date] = Query(default=None, alias="to"),
    project_id: Optional[str] = None,
    db: Session = Depends(get_db),
    _user: dict = Depends(require_agent),
):
    """실시간 통계: GitLab에서 직접 조회. 날짜 범위가 지정되면 해당 기간의 신규/완료 건수를 반환."""
    if from_date and to_date:
        if from_date > to_date:
            raise HTTPException(status_code=400, detail="종료일은 시작일 이후여야 합니다.")
        if (to_date - from_date) > timedelta(days=366):
            raise HTTPException(status_code=400, detail="날짜 범위는 최대 366일까지 허용됩니다.")

    # GitLab API 날짜 필터: ISO 형식 (날짜 경계를 명확히 하기 위해 datetime 문자열 사용)
    from_iso = datetime.combine(from_date, dt_time.min).isoformat() if from_date else None
    to_iso = datetime.combine(to_date, dt_time.max).isoformat() if to_date else None

    # Redis 캐시 조회
    import json as _json
    _r = get_redis()
    _cache_key = f"rpt:cs:{from_date}:{to_date}:{project_id or ''}"
    if _r:
        try:
            _cached = _r.get(_cache_key)
            if _cached:
                return _json.loads(_cached)
        except Exception:
            pass

    def _count_new():
        """기간 내 신규 생성된 티켓 수"""
        _, total = gitlab_client.get_issues(
            state="all",
            project_id=project_id,
            per_page=1, page=1,
            created_after=from_iso,
            created_before=to_iso,
        )
        return total

    def _count_open():
        """기간 내 생성된 티켓 중 현재 접수됨 상태"""
        _, total = gitlab_client.get_issues(
            state="opened",
            not_labels="status::in_progress,status::waiting,status::resolved",
            project_id=project_id,
            per_page=1, page=1,
            created_after=from_iso,
            created_before=to_iso,
        )
        return total

    def _count_in_progress():
        """기간 내 생성된 티켓 중 현재 처리 중"""
        _, total = gitlab_client.get_issues(
            state="opened",
            labels="status::in_progress",
            project_id=project_id,
            per_page=1, page=1,
            created_after=from_iso,
            created_before=to_iso,
        )
        return total

    def _count_resolved():
        """기간 내 생성된 티켓 중 현재 처리 완료"""
        _, total = gitlab_client.get_issues(
            state="opened",
            labels="status::resolved",
            project_id=project_id,
            per_page=1, page=1,
            created_after=from_iso,
            created_before=to_iso,
        )
        return total

    def _count_closed():
        """기간 내 종료된 티켓 수 (updated_after 근사치)"""
        _, total = gitlab_client.get_issues(
            state="closed",
            project_id=project_id,
            per_page=1, page=1,
            updated_after=from_iso,
            updated_before=to_iso,
        )
        return total

    def _count_sla_breached():
        """기간 내 SLA 위반 건수"""
        q = db.query(SLARecord).filter(SLARecord.breached == True)  # noqa: E712
        if from_date:
            q = q.filter(SLARecord.sla_deadline >= datetime.combine(from_date, dt_time.min))
        if to_date:
            q = q.filter(SLARecord.sla_deadline <= datetime.combine(to_date, dt_time.max))
        if project_id:
            q = q.filter(SLARecord.project_id == project_id)
        return q.count()

    try:
        # DB 세션은 스레드 안전하지 않으므로 ThreadPool 밖에서 먼저 조회
        sla_breached_count = _count_sla_breached()

        with ThreadPoolExecutor(max_workers=5) as pool:
            f_new         = pool.submit(_count_new)
            f_open        = pool.submit(_count_open)
            f_in_progress = pool.submit(_count_in_progress)
            f_resolved    = pool.submit(_count_resolved)
            f_closed      = pool.submit(_count_closed)

            result = {
                "new":         f_new.result(),
                "open":        f_open.result(),
                "in_progress": f_in_progress.result(),
                "resolved":    f_resolved.result(),
                "closed":      f_closed.result(),
                "sla_breached":sla_breached_count,
                "fetched_at":  datetime.now(timezone.utc).isoformat(),
            }

        if _r:
            try:
                _r.setex(_cache_key, _CACHE_TTL_STATS, _json.dumps(result))
            except Exception:
                pass
        return result
    except Exception as e:
        logger.error("current-stats GitLab error: %s", e)
        raise HTTPException(status_code=502, detail="통계를 불러오는 중 오류가 발생했습니다.")


@router.get("/trends")
def get_trends(
    from_date: Optional[date] = Query(default=None, alias="from"),
    to_date: Optional[date] = Query(default=None, alias="to"),
    project_id: Optional[str] = None,
    db: Session = Depends(get_db),
    _user: dict = Depends(require_agent),
):
    """날짜별 티켓 스냅샷 추이 (자정마다 기록된 일별 통계)."""
    if from_date and to_date:
        if from_date > to_date:
            raise HTTPException(status_code=400, detail="종료일은 시작일 이후여야 합니다.")
        if (to_date - from_date) > timedelta(days=366):
            raise HTTPException(status_code=400, detail="날짜 범위는 최대 366일까지 허용됩니다.")
    q = db.query(DailyStatsSnapshot).order_by(DailyStatsSnapshot.snapshot_date)
    if from_date:
        q = q.filter(DailyStatsSnapshot.snapshot_date >= from_date)
    if to_date:
        q = q.filter(DailyStatsSnapshot.snapshot_date <= to_date)
    if project_id:
        q = q.filter(DailyStatsSnapshot.project_id == project_id)

    rows = q.limit(3650).all()  # 최대 10년치 일별 스냅샷 (무제한 방지)
    return [
        {
            "snapshot_date":      str(r.snapshot_date),
            "project_id":         r.project_id,
            "total_open":         r.total_open,
            "total_in_progress":  r.total_in_progress,
            "total_closed":       r.total_closed,
            "total_new":          r.total_new,
            "total_breached":     r.total_breached,
        }
        for r in rows
    ]


@router.get("/export")
def export_report(
    from_date: Optional[date] = Query(default=None, alias="from"),
    to_date: Optional[date] = Query(default=None, alias="to"),
    project_id: Optional[str] = None,
    format: str = Query(default="csv"),
    db: Session = Depends(get_db),
    _user: dict = Depends(require_agent),
):
    if from_date and to_date:
        if from_date > to_date:
            raise HTTPException(status_code=400, detail="종료일은 시작일 이후여야 합니다.")
        if (to_date - from_date) > timedelta(days=366):
            raise HTTPException(status_code=400, detail="날짜 범위는 최대 366일까지 허용됩니다.")
    q = db.query(DailyStatsSnapshot).order_by(DailyStatsSnapshot.snapshot_date)
    if from_date:
        q = q.filter(DailyStatsSnapshot.snapshot_date >= from_date)
    if to_date:
        q = q.filter(DailyStatsSnapshot.snapshot_date <= to_date)
    if project_id:
        q = q.filter(DailyStatsSnapshot.project_id == project_id)

    rows = q.limit(3650).all()  # 최대 10년치 (내보내기용)
    date_str = datetime.now().strftime("%Y%m%d")

    if format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(status_code=501, detail="openpyxl 라이브러리가 설치되지 않았습니다.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ITSM 보고서"

        headers = ["날짜", "프로젝트", "당일 신규", "접수됨", "처리 중(대기·완료 포함)", "누적 종료", "SLA 위반 누적"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, r in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=str(r.snapshot_date))
            ws.cell(row=row_idx, column=2, value=r.project_id or "")
            ws.cell(row=row_idx, column=3, value=r.total_new)
            ws.cell(row=row_idx, column=4, value=r.total_open)
            ws.cell(row=row_idx, column=5, value=r.total_in_progress)
            ws.cell(row=row_idx, column=6, value=r.total_closed)
            ws.cell(row=row_idx, column=7, value=r.total_breached)

        # 컬럼 너비 자동 조정
        col_widths = [12, 20, 12, 10, 22, 12, 16]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        xlsx_buf = io.BytesIO()
        wb.save(xlsx_buf)
        xlsx_buf.seek(0)

        filename = f"itsm_report_{date_str}.xlsx"
        return StreamingResponse(
            iter([xlsx_buf.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
        )

    # 기본: CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["날짜", "프로젝트", "당일신규", "접수됨", "처리중(대기·완료포함)", "누적종료", "SLA위반누적"])
    for r in rows:
        writer.writerow([
            _sanitize_csv_cell(str(r.snapshot_date)),
            _sanitize_csv_cell(r.project_id or ""),
            r.total_new,
            r.total_open,
            r.total_in_progress,
            r.total_closed,
            r.total_breached,
        ])

    filename = f"itsm_report_{date_str}.csv"
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.get("/breakdown")
def get_breakdown(
    from_date: Optional[date] = Query(default=None, alias="from"),
    to_date: Optional[date] = Query(default=None, alias="to"),
    project_id: Optional[str] = None,
    db: Session = Depends(get_db),
    _user: dict = Depends(require_agent),
):
    """기간 내 티켓을 상태별·카테고리별·우선순위별로 집계."""
    if from_date and to_date:
        if from_date > to_date:
            raise HTTPException(status_code=400, detail="종료일은 시작일 이후여야 합니다.")
        if (to_date - from_date) > timedelta(days=366):
            raise HTTPException(status_code=400, detail="날짜 범위는 최대 366일까지 허용됩니다.")
    from_iso = datetime.combine(from_date, dt_time.min).isoformat() if from_date else None
    to_iso = datetime.combine(to_date, dt_time.max).isoformat() if to_date else None

    import json as _json
    _r = get_redis()
    _cache_key = f"rpt:bk:{from_date}:{to_date}:{project_id or ''}"
    if _r:
        try:
            _cached = _r.get(_cache_key)
            if _cached:
                return _json.loads(_cached)
        except Exception:
            pass

    all_issues: list[dict] = []
    page = 1
    while True:
        issues, total = gitlab_client.get_issues(
            state="all", per_page=100, page=page,
            project_id=project_id,
            created_after=from_iso, created_before=to_iso,
        )
        all_issues.extend(issues)
        if not issues or len(all_issues) >= total or page >= 50:  # 5,000건 안전 캡
            break
        page += 1

    by_status: dict[str, int] = {
        "open": 0, "in_progress": 0, "waiting": 0,
        "resolved": 0, "closed": 0,
    }
    by_category: dict[str, int] = {}
    by_priority: dict[str, int] = {}
    _VALID_PRIORITIES = {"low", "medium", "high", "critical"}
    _CAT_KO = {"hardware": "하드웨어", "software": "소프트웨어",
               "network": "네트워크", "account": "계정/권한", "other": "기타"}

    for issue in all_issues:
        labels = issue.get("labels", [])
        state = issue.get("state", "")

        if state == "closed":
            by_status["closed"] += 1
        else:
            status_lbl = next((l[8:] for l in labels if l.startswith("status::")), None)
            if status_lbl == "in_progress":
                by_status["in_progress"] += 1
            elif status_lbl == "waiting":
                by_status["waiting"] += 1
            elif status_lbl == "resolved":
                by_status["resolved"] += 1
            else:
                by_status["open"] += 1

        cat = next((l[5:] for l in labels if l.startswith("cat::")), "기타")
        cat = _CAT_KO.get(cat, cat)
        by_category[cat] = by_category.get(cat, 0) + 1

        prio = next((l[6:] for l in labels if l.startswith("prio::")), "medium")
        # corrupt 라벨 정규화: PriorityEnum.MEDIUM → medium, HIGH → high
        if "." in prio:
            prio = prio.split(".")[-1]
        prio = prio.lower()
        if prio not in _VALID_PRIORITIES:
            prio = "medium"
        by_priority[prio] = by_priority.get(prio, 0) + 1

    breakdown_result = {
        "total": len(all_issues),
        "by_status": by_status,
        "by_category": by_category,
        "by_priority": by_priority,
    }
    if _r:
        try:
            _r.setex(_cache_key, _CACHE_TTL_BREAKDOWN, _json.dumps(breakdown_result))
        except Exception:
            pass
    return breakdown_result


@router.get("/ratings")
def get_rating_stats(
    from_date: Optional[date] = Query(default=None, alias="from"),
    to_date: Optional[date] = Query(default=None, alias="to"),
    db: Session = Depends(get_db),
    _user: dict = Depends(require_agent),
):
    """만족도 평가 통계: 평균, 분포, 최근 평가 목록."""
    q = db.query(Rating).order_by(Rating.created_at.desc())
    if from_date:
        q = q.filter(Rating.created_at >= datetime.combine(from_date, dt_time.min))
    if to_date:
        q = q.filter(Rating.created_at <= datetime.combine(to_date, dt_time.max))
    ratings = q.limit(50000).all()  # 평가 집계용 최대 5만건
    total = len(ratings)

    if total == 0:
        return {
            "total": 0,
            "average": None,
            "distribution": {1: 0, 2: 0, 3: 0, 4: 0, 5: 0},
            "recent": [],
        }

    average = round(sum(r.score for r in ratings) / total, 2)
    distribution = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    for r in ratings:
        distribution[r.score] = distribution.get(r.score, 0) + 1

    recent = [
        {
            "id": r.id,
            "gitlab_issue_iid": r.gitlab_issue_iid,
            "employee_name": r.employee_name,
            "score": r.score,
            "comment": r.comment,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in ratings[:20]
    ]

    low_ratings = [
        {
            "id": r.id,
            "gitlab_issue_iid": r.gitlab_issue_iid,
            "employee_name": r.employee_name,
            "score": r.score,
            "comment": r.comment,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in ratings
        if r.score <= 2
    ][:20]

    return {
        "total": total,
        "average": average,
        "distribution": distribution,
        "recent": recent,
        "low_ratings": low_ratings,
    }


@router.get("/agent-performance")
def get_agent_performance(
    from_date: Optional[date] = Query(default=None, alias="from"),
    to_date: Optional[date] = Query(default=None, alias="to"),
    project_id: Optional[str] = None,
    db: Session = Depends(get_db),
    _user: dict = Depends(require_agent),
):
    """F-7: Per-agent performance stats: assigned, resolved, avg_rating, sla_met_rate."""
    import json as _json
    from_dt = datetime.combine(from_date, dt_time.min) if from_date else None
    to_dt = datetime.combine(to_date, dt_time.max) if to_date else None

    # Redis 캐시 조회
    _r = get_redis()
    _cache_key = f"rpt:ap:{from_date}:{to_date}:{project_id or ''}"
    if _r:
        try:
            _cached = _r.get(_cache_key)
            if _cached:
                return _json.loads(_cached)
        except Exception:
            pass

    # Gather all issues from GitLab with assignees
    all_issues: list[dict] = []
    try:
        page = 1
        from_iso = from_dt.isoformat() if from_dt else None
        to_iso = to_dt.isoformat() if to_dt else None
        while True:
            issues, total = gitlab_client.get_issues(
                state="all", per_page=100, page=page,
                project_id=project_id,
                created_after=from_iso,
                created_before=to_iso,
            )
            all_issues.extend(issues)
            if len(all_issues) >= total or not issues or page >= 50:  # 5,000건 안전 캡
                break
            page += 1
    except Exception as e:
        logger.error("agent-performance GitLab error: %s", e)

    # Aggregate by assignee
    agents: dict[str, dict] = {}
    for issue in all_issues:
        assignees = issue.get("assignees", [])
        if not assignees:
            continue
        assignee = assignees[0]
        agent_name = assignee.get("name", assignee.get("username", "unknown"))
        agent_username = assignee.get("username", "unknown")
        key = agent_username
        if key not in agents:
            agents[key] = {
                "agent_name": agent_name,
                "agent_username": agent_username,
                "assigned": 0,
                "resolved": 0,
                "sla_met": 0,
                "sla_total": 0,
            }
        agents[key]["assigned"] += 1
        if issue.get("state") == "closed":
            agents[key]["resolved"] += 1

    # SLA met rate from DB
    sla_q = db.query(SLARecord)
    if project_id:
        sla_q = sla_q.filter(SLARecord.project_id == project_id)
    if from_dt:
        sla_q = sla_q.filter(SLARecord.created_at >= from_dt)
    if to_dt:
        sla_q = sla_q.filter(SLARecord.created_at <= to_dt)
    sla_records = sla_q.limit(50000).all()

    sla_iid_map = {r.gitlab_issue_iid: r for r in sla_records}
    for issue in all_issues:
        iid = issue.get("iid")
        assignees = issue.get("assignees", [])
        if not assignees or iid not in sla_iid_map:
            continue
        key = assignees[0].get("username", "unknown")
        if key not in agents:
            continue
        agents[key]["sla_total"] += 1
        if not sla_iid_map[iid].breached:
            agents[key]["sla_met"] += 1

    # Ratings per issue → map to assignee
    rating_q = db.query(Rating)
    if from_dt:
        rating_q = rating_q.filter(Rating.created_at >= from_dt)
    if to_dt:
        rating_q = rating_q.filter(Rating.created_at <= to_dt)
    ratings = rating_q.limit(50000).all()
    rating_map: dict[int, list[int]] = {}
    for r in ratings:
        rating_map.setdefault(r.gitlab_issue_iid, []).append(r.score)

    iid_to_agent: dict[int, str] = {}
    for issue in all_issues:
        assignees = issue.get("assignees", [])
        if assignees:
            iid_to_agent[issue["iid"]] = assignees[0].get("username", "unknown")

    for iid, scores in rating_map.items():
        key = iid_to_agent.get(iid)
        if key and key in agents:
            agents[key].setdefault("_scores", []).extend(scores)

    result = []
    for data in agents.values():
        scores = data.pop("_scores", [])
        avg_rating = round(sum(scores) / len(scores), 2) if scores else None
        sla_met_rate = (
            round(data["sla_met"] / data["sla_total"] * 100, 1)
            if data["sla_total"] > 0 else None
        )
        result.append({
            "agent_name": data["agent_name"],
            "agent_username": data["agent_username"],
            "assigned": data["assigned"],
            "resolved": data["resolved"],
            "avg_rating": avg_rating,
            "sla_met_rate": sla_met_rate,
        })

    result.sort(key=lambda x: x["assigned"], reverse=True)

    # Redis 캐시 저장
    if _r:
        try:
            _r.setex(_cache_key, _CACHE_TTL_AGENT_PERF, _json.dumps(result))
        except Exception:
            pass

    return result


def take_snapshot(project_id: str, db) -> dict:
    """스냅샷 생성 내부 함수 — HTTP 컨텍스트 없이 호출 가능."""
    today = date.today()
    today_start = datetime.combine(today, dt_time.min).isoformat()
    today_end = datetime.combine(today, dt_time.max).isoformat()

    existing = db.query(DailyStatsSnapshot).filter(
        DailyStatsSnapshot.snapshot_date == today,
        DailyStatsSnapshot.project_id == project_id,
    ).first()
    if existing:
        return {"message": "already_exists", "date": str(today)}

    try:
        def _get(state, labels=None, not_labels=None, created_after=None, created_before=None):
            _, total = gitlab_client.get_issues(
                state=state, labels=labels, not_labels=not_labels,
                per_page=1, page=1, project_id=project_id,
                created_after=created_after, created_before=created_before,
            )
            return total

        # DB 세션은 스레드 안전하지 않으므로 ThreadPool 밖에서 먼저 조회
        total_breached = db.query(SLARecord).filter(
            SLARecord.project_id == project_id,
            SLARecord.breached == True,  # noqa: E712
        ).count()

        with ThreadPoolExecutor(max_workers=6) as pool:
            f_open     = pool.submit(_get, "opened", None, "status::in_progress,status::waiting,status::resolved")
            f_closed   = pool.submit(_get, "closed")
            f_ip       = pool.submit(_get, "opened", "status::in_progress")
            f_waiting  = pool.submit(_get, "opened", "status::waiting")
            f_resolved = pool.submit(_get, "opened", "status::resolved")
            f_new      = pool.submit(_get, "all", None, None, today_start, today_end)
            total_open        = f_open.result()
            total_closed      = f_closed.result()
            total_in_progress = f_ip.result() + f_waiting.result() + f_resolved.result()
            total_new         = f_new.result()
    except Exception as e:
        logger.error("Snapshot fetch error for project %s: %s", project_id, e)
        total_open = total_closed = total_in_progress = total_new = total_breached = 0

    snapshot = DailyStatsSnapshot(
        snapshot_date=today,
        project_id=project_id,
        total_open=total_open,
        total_in_progress=total_in_progress,
        total_closed=total_closed,
        total_new=total_new,
        total_breached=total_breached,
    )
    db.add(snapshot)
    try:
        db.commit()
    except IntegrityError:
        # 동시 실행(스케줄러 + on-access)으로 인한 중복 → 무시
        db.rollback()
        return {"message": "already_exists", "date": str(today)}
    return {"message": "created", "date": str(today)}


@router.post("/snapshots", status_code=201)
def create_daily_snapshot(
    project_id: str,
    db: Session = Depends(get_db),
    _user: dict = Depends(require_agent),
):
    """Manually trigger a daily stats snapshot for a project."""
    result = take_snapshot(project_id, db)
    if result["message"] == "already_exists":
        return {"message": "오늘 스냅샷이 이미 있습니다.", "date": result["date"]}
    return {"message": "스냅샷이 생성됐습니다.", "date": result["date"]}


@router.get("/dora")
def get_dora_metrics(
    days: int = Query(default=30, ge=7, le=365),
    project_id: Optional[str] = None,
    db: Session = Depends(get_db),
    _user: dict = Depends(require_agent),
):
    """DORA 4대 지표 반환.

    - deployment_frequency: 기간 내 completed(closed) 티켓 수 / 주 (변경 배포 빈도 근사치)
    - lead_time_hours: 접수 → 완료까지 평균 리드타임 (시간)
    - change_failure_rate: 완료 후 재오픈된 비율 (%)
    - mttr_hours: 재오픈된 티켓의 두 번째 완료까지 평균 시간 (시간)
    """
    now = datetime.now(timezone.utc)
    since = (now - timedelta(days=days)).replace(tzinfo=None)
    since_iso = since.isoformat()

    pid = project_id or None

    # ── 1. Deployment Frequency ──────────────────────────────────────────
    # closed 티켓 수를 주(week) 단위로 나눔
    try:
        _, total_closed = gitlab_client.get_issues(
            state="closed",
            project_id=pid,
            per_page=1, page=1,
            updated_after=since_iso,
        )
    except Exception:
        total_closed = 0
    weeks = max(days / 7, 1)
    deployment_frequency = round(total_closed / weeks, 2)

    # ── 2. Lead Time (접수 → 완료 평균, SLARecord 기반) ──────────────────
    try:
        records = (
            db.query(SLARecord)
            .filter(
                SLARecord.resolved_at.isnot(None),
                SLARecord.created_at >= since,
            )
        )
        if pid:
            records = records.filter(SLARecord.project_id == pid)
        records = records.all()

        if records:
            lead_times_h = [
                (r.resolved_at - r.created_at).total_seconds() / 3600
                for r in records
                if r.resolved_at and r.created_at
            ]
            lead_time_hours = round(sum(lead_times_h) / len(lead_times_h), 1) if lead_times_h else None
        else:
            lead_time_hours = None
    except Exception as e:
        logger.warning("DORA lead_time query error: %s", e)
        lead_time_hours = None

    # ── 3. Change Failure Rate (완료 후 재오픈 비율) ──────────────────────
    # SLA record 중 resolved_at가 있고 이후 재접수된(sla_deadline 갱신) 티켓
    try:
        total_resolved = (
            db.query(SLARecord)
            .filter(
                SLARecord.resolved_at.isnot(None),
                SLARecord.created_at >= since,
            )
        )
        reopened = (
            db.query(SLARecord)
            .filter(
                SLARecord.resolved_at.isnot(None),
                SLARecord.reopened_at.isnot(None),
                SLARecord.created_at >= since,
            )
        )
        if pid:
            total_resolved = total_resolved.filter(SLARecord.project_id == pid)
            reopened = reopened.filter(SLARecord.project_id == pid)

        total_r = total_resolved.count()
        reopened_r = reopened.count()
        change_failure_rate = round(reopened_r / total_r * 100, 1) if total_r > 0 else 0.0
    except Exception as e:
        logger.warning("DORA change_failure_rate query error: %s", e)
        change_failure_rate = 0.0
        reopened_r = 0
        total_r = 0

    # ── 4. MTTR (재오픈된 티켓의 두 번째 완료까지 평균 시간) ──────────────
    try:
        reopened_records = (
            db.query(SLARecord)
            .filter(
                SLARecord.resolved_at.isnot(None),
                SLARecord.reopened_at.isnot(None),
                SLARecord.created_at >= since,
            )
        )
        if pid:
            reopened_records = reopened_records.filter(SLARecord.project_id == pid)
        reopened_records = reopened_records.all()

        if reopened_records:
            mttr_values = [
                (r.resolved_at - r.reopened_at).total_seconds() / 3600
                for r in reopened_records
                if r.resolved_at and r.reopened_at and r.resolved_at > r.reopened_at
            ]
            mttr_hours = round(sum(mttr_values) / len(mttr_values), 1) if mttr_values else None
        else:
            mttr_hours = None
    except Exception as e:
        logger.warning("DORA mttr query error: %s", e)
        mttr_hours = None

    # ── 등급 계산 (Elite / High / Medium / Low) ────────────────────────────
    def _df_grade(df: float) -> str:
        if df >= 7:    return "Elite"
        if df >= 1:    return "High"
        if df >= 0.25: return "Medium"
        return "Low"

    def _lt_grade(lt: Optional[float]) -> str:
        if lt is None: return "N/A"
        if lt <= 24:   return "Elite"
        if lt <= 168:  return "High"
        if lt <= 720:  return "Medium"
        return "Low"

    def _cfr_grade(cfr: float) -> str:
        if cfr <= 5:  return "Elite"
        if cfr <= 10: return "High"
        if cfr <= 15: return "Medium"
        return "Low"

    def _mttr_grade(mt: Optional[float]) -> str:
        if mt is None:  return "N/A"
        if mt <= 1:     return "Elite"
        if mt <= 24:    return "High"
        if mt <= 168:   return "Medium"
        return "Low"

    return {
        "period_days": days,
        "since": since_iso,
        "deployment_frequency": {
            "value": deployment_frequency,
            "unit": "회/주",
            "grade": _df_grade(deployment_frequency),
            "description": "주간 완료(closed) 티켓 수",
        },
        "lead_time": {
            "value": lead_time_hours,
            "unit": "시간",
            "grade": _lt_grade(lead_time_hours),
            "description": "접수에서 완료까지 평균 리드타임",
        },
        "change_failure_rate": {
            "value": change_failure_rate,
            "unit": "%",
            "grade": _cfr_grade(change_failure_rate),
            "description": "완료 후 재오픈된 티켓 비율",
            "resolved_count": total_r,
            "reopened_count": reopened_r,
        },
        "mttr": {
            "value": mttr_hours,
            "unit": "시간",
            "grade": _mttr_grade(mttr_hours),
            "description": "재오픈 후 재완료까지 평균 복구 시간",
        },
    }


@router.get("/sla-dashboard")
def get_sla_dashboard(
    project_id: Optional[str] = None,
    db: Session = Depends(get_db),
    _user: dict = Depends(require_agent),
):
    """SLA 에스컬레이션 대시보드.

    위반/임박/정상 티켓 수와 함께 위반·임박 티켓 목록, 최근 7일 위반 트렌드를 반환한다.
    """
    import json as _json
    _r = get_redis()
    _cache_key = f"rpt:sla_dash:{project_id or ''}"
    if _r:
        try:
            _cached = _r.get(_cache_key)
            if _cached:
                return _json.loads(_cached)
        except Exception:
            pass

    now = datetime.now(timezone.utc).replace(tzinfo=None)

    base_q = db.query(SLARecord).filter(
        SLARecord.resolved_at == None,  # noqa: E711
        SLARecord.paused_at == None,    # noqa: E711
    )
    if project_id:
        base_q = base_q.filter(SLARecord.project_id == project_id)

    active_records: list[SLARecord] = base_q.limit(10000).all()

    breach_list = []
    warning_list = []
    on_track_count = 0

    for rec in active_records:
        deadline = rec.sla_deadline
        if deadline is None:
            continue
        total_seconds = (now - rec.created_at).total_seconds() if rec.created_at else 1
        sla_window = (deadline - rec.created_at).total_seconds() if rec.created_at else 1
        elapsed_pct = round((total_seconds / sla_window) * 100, 1) if sla_window > 0 else 0

        remaining_seconds = round((deadline - now).total_seconds())
        ticket_data = {
            "iid": rec.gitlab_issue_iid,
            "project_id": rec.project_id,
            "priority": rec.priority,
            "sla_deadline": deadline.isoformat(),
            "elapsed_pct": elapsed_pct,
            "remaining_seconds": remaining_seconds,
            "breached": rec.breached,
        }

        if rec.breached or now >= deadline:
            breach_list.append(ticket_data)
        elif elapsed_pct >= 80:
            warning_list.append(ticket_data)
        else:
            on_track_count += 1

    # GitLab에서 위반/임박 티켓 상세 정보 병렬 조회
    combined = breach_list + warning_list

    def _enrich_one(item: dict) -> dict:
        try:
            issue = gitlab_client.get_issue(item["iid"], project_id=item["project_id"])
            assignee = None
            if issue.get("assignee"):
                assignee = issue["assignee"].get("name") or issue["assignee"].get("username")
            elif issue.get("assignees"):
                first = issue["assignees"][0]
                assignee = first.get("name") or first.get("username")
            return {
                "iid": item["iid"],
                "title": issue.get("title", f"티켓 #{item['iid']}"),
                "status": next(
                    (lb.split("::")[1] for lb in issue.get("labels", []) if lb.startswith("status::")),
                    "open",
                ),
                "priority": item["priority"],
                "sla_deadline": item["sla_deadline"],
                "elapsed_pct": item["elapsed_pct"],
                "remaining_seconds": item["remaining_seconds"],
                "assignee": assignee,
                "breached": item["breached"],
            }
        except Exception as e:
            logger.warning("SLA dashboard: failed to fetch issue #%s: %s", item["iid"], e)
            # 404 또는 기타 오류 시 None 반환 → 목록에서 제외
            return None

    enriched_tickets: list[dict] = []
    if combined:
        max_workers = min(len(combined), 10)
        with ThreadPoolExecutor(max_workers=max_workers) as _pool:
            enriched_tickets = [t for t in _pool.map(_enrich_one, combined) if t is not None]

    # breach 먼저, 그 다음 elapsed_pct 내림차순 정렬
    enriched_tickets.sort(key=lambda t: (0 if t["breached"] else 1, -t["elapsed_pct"]))

    # 최근 7일 일별 위반 건수
    today_d = date.today()
    seven_days_ago = today_d - timedelta(days=6)
    trend_rows = (
        db.query(DailyStatsSnapshot)
        .filter(DailyStatsSnapshot.snapshot_date >= seven_days_ago)
        .order_by(DailyStatsSnapshot.snapshot_date)
        .all()
    )
    from collections import defaultdict
    by_date: dict[str, int] = defaultdict(int)
    for r in trend_rows:
        by_date[str(r.snapshot_date)] += r.total_breached or 0

    trend = []
    cursor = seven_days_ago
    while cursor <= today_d:
        key = str(cursor)
        trend.append({"date": key, "count": by_date[key]})
        cursor += timedelta(days=1)

    # enriched 목록 기준으로 카운트 (orphan SLA 레코드 제외)
    enriched_breach = sum(1 for t in enriched_tickets if t["breached"])
    enriched_warning = len(enriched_tickets) - enriched_breach
    sla_dash_result = {
        "breach_count": enriched_breach,
        "warning_count": enriched_warning,
        "on_track_count": on_track_count,
        "tickets": enriched_tickets,
        "trend": trend,
    }
    if _r:
        try:
            _r.setex(_cache_key, _CACHE_TTL_SLA_DASH, _json.dumps(sla_dash_result))
        except Exception:
            pass
    return sla_dash_result


@router.get("/sla/heatmap")
def get_sla_heatmap(
    weeks: int = Query(default=12, ge=4, le=52),
    db: Session = Depends(get_db),
    _user: dict = Depends(require_agent),
):
    """주차별 SLA 위반 히트맵 데이터 반환.

    최근 N주 동안 각 날짜의 SLA 위반 건수를 반환한다.
    응답: [{"date": "YYYY-MM-DD", "breached": int, "total": int}, ...]
    """
    today_d = date.today()
    # 월요일 기준으로 weeks * 7일 전으로 맞춤
    start_dow = today_d.weekday()  # 0=월 … 6=일
    # 이번 주 월요일
    this_monday = today_d - timedelta(days=start_dow)
    since = this_monday - timedelta(weeks=weeks - 1)

    rows = (
        db.query(DailyStatsSnapshot)
        .filter(DailyStatsSnapshot.snapshot_date >= since)
        .order_by(DailyStatsSnapshot.snapshot_date)
        .all()
    )

    # 날짜별로 여러 프로젝트 행을 합산
    from collections import defaultdict
    by_date: dict[str, dict] = defaultdict(lambda: {"breached": 0, "total": 0})
    for r in rows:
        key = str(r.snapshot_date)
        by_date[key]["breached"] += r.total_breached or 0
        by_date[key]["total"] += (r.total_open or 0) + (r.total_in_progress or 0) + (r.total_closed or 0)

    # since ~ today 범위 내 모든 날짜를 채움 (스냅샷 없는 날은 0)
    result = []
    cursor = since
    while cursor <= today_d:
        key = str(cursor)
        result.append({
            "date": key,
            "breached": by_date[key]["breached"],
            "total": by_date[key]["total"],
        })
        cursor += timedelta(days=1)
    return result


@router.get("/csat-trend")
def get_csat_trend(
    from_date: Optional[date] = Query(default=None, alias="from"),
    to_date: Optional[date] = Query(default=None, alias="to"),
    granularity: str = Query(default="weekly", pattern="^(weekly|monthly)$"),
    db: Session = Depends(get_db),
    _user: dict = Depends(require_agent),
):
    """CSAT 트렌드: 주별 또는 월별로 만족도 통계를 집계합니다.

    - granularity=weekly : ISO 연도-주(YYYY-Www) 기준
    - granularity=monthly: YYYY-MM 기준
    csat_pct = 4점 이상 비율(%)
    """
    if from_date and to_date and from_date > to_date:
        raise HTTPException(status_code=400, detail="종료일은 시작일 이후여야 합니다.")

    q = db.query(Rating).order_by(Rating.created_at)
    if from_date:
        q = q.filter(Rating.created_at >= datetime.combine(from_date, dt_time.min))
    if to_date:
        q = q.filter(Rating.created_at <= datetime.combine(to_date, dt_time.max))
    ratings = q.all()

    buckets: dict[str, list[int]] = {}
    for r in ratings:
        dt = r.created_at
        if dt is None:
            continue
        if granularity == "monthly":
            key = dt.strftime("%Y-%m")
        else:
            iso_cal = dt.isocalendar()
            key = f"{iso_cal[0]}-W{iso_cal[1]:02d}"
        buckets.setdefault(key, []).append(r.score)

    result = []
    for period in sorted(buckets.keys()):
        scores = buckets[period]
        count = len(scores)
        average = round(sum(scores) / count, 2) if count else None
        csat_count = sum(1 for s in scores if s >= 4)
        csat_pct = round(csat_count / count * 100, 1) if count else None
        result.append({
            "period": period,
            "count": count,
            "average": average,
            "csat_pct": csat_pct,
        })
    return result


# ---------------------------------------------------------------------------
# Task 4: 시간 추적 리포트
# ---------------------------------------------------------------------------

@router.get("/time-tracking")
def get_time_tracking_report(
    project_id: Optional[str] = None,
    agent: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_agent),
):
    """팀원별·티켓별 시간 기록 집계."""
    q = db.query(TimeEntry)
    if project_id:
        q = q.filter(TimeEntry.project_id == project_id)
    if agent:
        q = q.filter(TimeEntry.agent_id == agent)
    if start:
        try:
            start_dt = datetime.fromisoformat(start)
            q = q.filter(TimeEntry.logged_at >= start_dt)
        except ValueError:
            pass
    if end:
        try:
            end_dt = datetime.fromisoformat(end) + timedelta(days=1)
            q = q.filter(TimeEntry.logged_at < end_dt)
        except ValueError:
            pass

    entries = q.order_by(TimeEntry.logged_at.desc()).all()

    # per-agent aggregation
    agent_map: dict[str, dict] = {}
    for e in entries:
        key = e.agent_id
        if key not in agent_map:
            agent_map[key] = {
                "agent_id": e.agent_id,
                "agent_name": e.agent_name,
                "total_minutes": 0,
                "ticket_count": 0,
                "tickets": set(),
            }
        agent_map[key]["total_minutes"] += e.minutes
        agent_map[key]["tickets"].add(e.issue_iid)

    by_agent = []
    for a in sorted(agent_map.values(), key=lambda x: -x["total_minutes"]):
        by_agent.append({
            "agent_id": a["agent_id"],
            "agent_name": a["agent_name"],
            "total_minutes": a["total_minutes"],
            "total_hours": round(a["total_minutes"] / 60, 1),
            "ticket_count": len(a["tickets"]),
        })

    # per-date aggregation
    date_map: dict[str, int] = {}
    for e in entries:
        key = e.logged_at.strftime("%Y-%m-%d") if e.logged_at else "unknown"
        date_map[key] = date_map.get(key, 0) + e.minutes
    by_date = [{"date": k, "minutes": v} for k, v in sorted(date_map.items())]

    # recent entries
    recent = [
        {
            "id": e.id,
            "issue_iid": e.issue_iid,
            "project_id": e.project_id,
            "agent_id": e.agent_id,
            "agent_name": e.agent_name,
            "minutes": e.minutes,
            "description": e.description,
            "logged_at": e.logged_at.isoformat() if e.logged_at else None,
        }
        for e in entries[:50]
    ]

    total_minutes = sum(e.minutes for e in entries)
    return {
        "total_minutes": total_minutes,
        "total_hours": round(total_minutes / 60, 1),
        "entry_count": len(entries),
        "agent_count": len(by_agent),
        "by_agent": by_agent,
        "by_date": by_date,
        "recent_entries": recent,
    }


# ---------------------------------------------------------------------------
# Task 6: 멀티 프로젝트 통합 뷰
# ---------------------------------------------------------------------------

@router.get("/multi-project")
def get_multi_project_stats(
    db: Session = Depends(get_db),
    user: dict = Depends(require_agent),
):
    """등록된 모든 GitLab 프로젝트의 SLA·티켓 현황을 통합 조회한다."""
    from ..models import SLARecord

    # 사용된 project_id 목록 (SLA 기록 기준)
    rows = db.query(SLARecord.project_id).distinct().all()
    project_ids = [r[0] for r in rows if r[0]]

    # GitLab 프로젝트 이름 캐시
    project_names: dict[str, str] = {}
    try:
        projects = gitlab_client.get_user_accessible_projects(
            user.get("gitlab_token", "")
        )
        for p in projects:
            project_names[str(p.get("id", ""))] = p.get("name", "")
    except Exception:
        pass

    # SQL 집계로 N+1 제거 — 프로젝트별 SLA 통계를 한 번의 쿼리로 조회
    from sqlalchemy import func, case
    sla_agg = (
        db.query(
            SLARecord.project_id,
            func.count().label("total"),
            func.sum(case((SLARecord.breached == True, 1), else_=0)).label("breached"),  # noqa: E712
            func.sum(case((SLARecord.resolved_at == None, 1), else_=0) * case((SLARecord.breached == False, 1), else_=0)).label("active"),  # noqa: E711,E712
        )
        .group_by(SLARecord.project_id)
        .all()
    )
    sla_map = {r.project_id: {"total": r.total, "breached": int(r.breached or 0), "active": int(r.active or 0)} for r in sla_agg}

    time_agg = (
        db.query(TimeEntry.project_id, func.coalesce(func.sum(TimeEntry.minutes), 0).label("total_min"))
        .group_by(TimeEntry.project_id)
        .all()
    )
    time_map = {r.project_id: int(r.total_min) for r in time_agg}

    result = []
    for pid in project_ids:
        s = sla_map.get(pid, {"total": 0, "breached": 0, "active": 0})
        total = s["total"]
        breached = s["breached"]
        result.append({
            "project_id": pid,
            "project_name": project_names.get(pid, pid),
            "total_sla_records": total,
            "sla_breached": breached,
            "sla_active": s["active"],
            "sla_compliance_rate": round((total - breached) / total * 100, 1) if total else None,
            "total_time_hours": round(time_map.get(pid, 0) / 60, 1),
        })

    result.sort(key=lambda x: -x["total_sla_records"])
    return {"projects": result}


# ---------------------------------------------------------------------------
# Task 9: SLA 준수율 트렌드 리포트
# ---------------------------------------------------------------------------

@router.get("/sla-compliance")
def get_sla_compliance_report(
    project_id: Optional[str] = None,
    weeks: int = Query(12, ge=1, le=52),
    db: Session = Depends(get_db),
    user: dict = Depends(require_agent),
):
    """주별 SLA 준수율 트렌드를 반환한다."""
    from ..models import SLARecord

    end_date = date.today()
    start_date = end_date - timedelta(weeks=weeks)
    start_dt = datetime.combine(start_date, dt_time.min)

    q = db.query(SLARecord).filter(SLARecord.created_at >= start_dt)
    if project_id:
        q = q.filter(SLARecord.project_id == project_id)
    records = q.all()

    # 주별 집계
    week_map: dict[str, dict] = {}
    for r in records:
        created = r.created_at
        if not created:
            continue
        # ISO 주 시작일 (월요일)
        iso_date = created.date()
        dow = iso_date.weekday()  # Mon=0
        week_start = iso_date - timedelta(days=dow)
        key = week_start.isoformat()
        if key not in week_map:
            week_map[key] = {"week": key, "total": 0, "met": 0, "breached": 0}
        week_map[key]["total"] += 1
        if r.breached:
            week_map[key]["breached"] += 1
        else:
            week_map[key]["met"] += 1

    trend = []
    for wk in sorted(week_map.keys()):
        d = week_map[wk]
        rate = round(d["met"] / d["total"] * 100, 1) if d["total"] else None
        trend.append({**d, "compliance_rate": rate})

    # 우선순위별 통계
    priority_map: dict[str, dict] = {}
    for r in records:
        p = r.priority or "unknown"
        if p not in priority_map:
            priority_map[p] = {"priority": p, "total": 0, "breached": 0}
        priority_map[p]["total"] += 1
        if r.breached:
            priority_map[p]["breached"] += 1
    by_priority = []
    for p, d in sorted(priority_map.items()):
        rate = round((d["total"] - d["breached"]) / d["total"] * 100, 1) if d["total"] else None
        by_priority.append({**d, "compliance_rate": rate})

    total = len(records)
    breached = sum(1 for r in records if r.breached)
    return {
        "period_weeks": weeks,
        "total": total,
        "met": total - breached,
        "breached": breached,
        "overall_compliance_rate": round((total - breached) / total * 100, 1) if total else None,
        "trend": trend,
        "by_priority": by_priority,
    }
