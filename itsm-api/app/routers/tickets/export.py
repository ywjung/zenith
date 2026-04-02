"""CSV and XLSX export/import endpoints."""
import csv
import io
import logging
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session

from ...auth import get_current_user
from ...database import get_db
from ... import gitlab_client
from ...rbac import require_pl, require_agent
from .helpers import _issue_to_response, _attach_sla_deadlines

logger = logging.getLogger(__name__)

export_router = APIRouter()


@export_router.get("/export/csv")
def export_tickets_csv(
    state: Optional[str] = Query(default=None),
    category: Optional[str] = Query(default=None),
    priority: Optional[str] = Query(default=None),
    search: Optional[str] = Query(default=None),
    project_id: Optional[str] = Query(default=None),
    _user: dict = Depends(require_pl),
):
    """현재 필터 기준 티켓 목록을 CSV로 내보낸다 (pl 이상)."""
    import csv
    import io
    from fastapi.responses import StreamingResponse as _StreamingResponse
    from datetime import date as _date

    labels = []
    if state and state != "all":
        if state == "closed":
            gl_state = "closed"
        else:
            labels.append(f"status::{state}")
            gl_state = "opened"
    else:
        gl_state = "all"
    if category:
        labels.append(f"cat::{category}")
    if priority:
        labels.append(f"prio::{priority}")

    label_str = ",".join(labels) if labels else None
    issues = gitlab_client.get_all_issues(
        state=gl_state, labels=label_str, search=search,
        project_id=project_id,
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["번호", "제목", "상태", "우선순위", "카테고리", "신청자", "담당자", "생성일", "수정일"])

    def _sc(v) -> str:
        """HIGH-06: CSV formula injection 방어 — reports.py의 _sanitize_csv_cell 동일 로직."""
        s = str(v) if v is not None else ""
        if s and s[0] in ('=', '+', '-', '@', '\t', '\r'):
            return "'" + s
        return s

    for issue in issues:
        ticket = _issue_to_response(issue)
        writer.writerow([
            ticket.get("iid"),
            _sc(ticket.get("title")),
            _sc(ticket.get("status")),
            _sc(ticket.get("priority")),
            _sc(ticket.get("category")),
            _sc(ticket.get("employee_name")),
            _sc(ticket.get("assignee_name")),
            ticket.get("created_at", "")[:10] if ticket.get("created_at") else "",
            ticket.get("updated_at", "")[:10] if ticket.get("updated_at") else "",
        ])

    output.seek(0)
    filename = f"tickets_{_date.today().isoformat()}.csv"
    return _StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@export_router.get("/export/xlsx")
def export_tickets_xlsx(
    state: Optional[str] = Query(default=None),
    category: Optional[str] = Query(default=None),
    priority: Optional[str] = Query(default=None),
    search: Optional[str] = Query(default=None),
    project_id: Optional[str] = Query(default=None),
    _user: dict = Depends(require_pl),
    db: Session = Depends(get_db),
):
    """현재 필터 기준 티켓 목록을 Excel(xlsx)로 내보낸다."""
    import io
    from datetime import date as _date
    from fastapi.responses import Response as _Response
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=501, detail="openpyxl 라이브러리가 설치되어 있지 않습니다.")

    labels = []
    if state and state != "all":
        if state == "closed":
            gl_state = "closed"
        else:
            labels.append(f"status::{state}")
            gl_state = "opened"
    else:
        gl_state = "all"
    if category:
        labels.append(f"cat::{category}")
    if priority:
        labels.append(f"prio::{priority}")

    label_str = ",".join(labels) if labels else None
    issues = gitlab_client.get_all_issues(
        state=gl_state, labels=label_str, search=search,
        project_id=project_id,
    )
    tickets = [_issue_to_response(i) for i in issues]
    _attach_sla_deadlines(tickets, db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "티켓 목록"

    headers = ["번호", "제목", "상태", "우선순위", "카테고리", "신청자", "담당자", "SLA 마감", "생성일", "수정일"]
    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    col_widths = [8, 45, 12, 10, 14, 16, 16, 14, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    priority_colors = {"critical": "FEE2E2", "high": "FEF3C7", "medium": "DBEAFE", "low": "F3F4F6"}
    for row_idx, t in enumerate(tickets, 2):
        prio = t.get("priority", "")
        row_fill = PatternFill(
            start_color=priority_colors.get(prio, "FFFFFF"),
            end_color=priority_colors.get(prio, "FFFFFF"),
            fill_type="solid",
        )
        values = [
            t.get("iid"),
            t.get("title", ""),
            t.get("status", ""),
            prio,
            t.get("category", ""),
            t.get("employee_name", ""),
            t.get("assignee_name", ""),
            (t.get("sla_deadline") or "")[:10],
            (t.get("created_at") or "")[:10],
            (t.get("updated_at") or "")[:10],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"tickets_{_date.today().isoformat()}.xlsx"
    return _Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@export_router.get("/import/template")
def download_import_template():
    """CSV 가져오기 템플릿 파일을 다운로드한다 (샘플 데이터 2행 포함)."""
    from fastapi.responses import Response as _Response

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["title", "description", "priority", "category", "assignee_username"])
    writer.writerow(["VPN 접속 오류", "재택근무 중 VPN 연결이 되지 않습니다.", "high", "network", ""])
    writer.writerow(["노트북 교체 요청", "노트북 배터리 수명이 다하여 교체가 필요합니다.", "medium", "hardware", "jdoe"])
    output.seek(0)
    content = "\ufeff" + output.getvalue()  # BOM for Excel compatibility
    return _Response(
        content=content.encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="itsm_import_template.csv"'},
    )


@export_router.post("/import/csv", status_code=201)
async def import_tickets_csv(
    file: UploadFile = File(...),
    project_id: Optional[str] = Query(default=None),
    _user: dict = Depends(require_agent),
    db: Session = Depends(get_db),
):
    """CSV 파일로 티켓 일괄 생성.

    필수 컬럼: title
    선택 컬럼: description, priority (critical/high/medium/low, 기본=medium),
              category (서비스 유형 값), assignee_username
    최대 200행.
    응답: {"imported": N, "failed": [{"row": N, "title": "...", "error": "..."}]}
    """
    from ...config import get_settings

    _VALID_PRIORITIES = {"critical", "high", "medium", "low"}
    _MAX_CSV_SIZE = 5 * 1024 * 1024  # 5MB — M3: 파일 크기 제한으로 메모리 DoS 방지

    # content-type 검증
    ct = (file.content_type or "").lower()
    if ct and ct not in {"text/csv", "application/csv", "text/plain", "application/octet-stream"}:
        raise HTTPException(status_code=422, detail="CSV 파일만 업로드 가능합니다.")

    content = await file.read()
    if len(content) > _MAX_CSV_SIZE:
        raise HTTPException(status_code=422, detail="CSV 파일은 최대 5MB까지 허용됩니다.")
    try:
        text = content.decode("utf-8-sig")  # BOM 제거
    except UnicodeDecodeError:
        text = content.decode("cp949", errors="replace")

    reader = csv.DictReader(io.StringIO(text))
    fieldnames = set(reader.fieldnames or [])
    if "title" not in fieldnames:
        raise HTTPException(status_code=422, detail="필수 컬럼 'title'이 없습니다.")

    rows = list(reader)
    if len(rows) > 200:
        raise HTTPException(status_code=422, detail="한 번에 최대 200행까지 가져올 수 있습니다.")

    pid = project_id or get_settings().GITLAB_PROJECT_ID
    imported = 0
    failed = []

    for idx, row in enumerate(rows, start=2):  # 1행 = 헤더
        title = (row.get("title") or "").strip()
        if not title:
            failed.append({"row": idx, "title": "", "error": "title이 비어있습니다."})
            continue

        description = (row.get("description") or "").strip()
        priority = (row.get("priority") or "medium").strip().lower()
        if priority not in _VALID_PRIORITIES:
            priority = "medium"
        category = (row.get("category") or "").strip().lower()
        assignee_username = (row.get("assignee_username") or "").strip()

        try:
            labels = ["status::open", f"prio::{priority}"]
            if category:
                labels.append(f"cat::{category}")

            # assignee_username → GitLab user_id 조회
            assignee_id: Optional[int] = None
            if assignee_username:
                try:
                    from ...config import get_settings as _gs
                    import httpx as _httpx
                    settings = _gs()
                    with _httpx.Client(timeout=10.0) as c:
                        resp = c.get(
                            f"{settings.GITLAB_API_URL}/api/v4/users",
                            headers={"PRIVATE-TOKEN": settings.GITLAB_PROJECT_TOKEN},
                            params={"username": assignee_username},
                        )
                        if resp.is_success:
                            users = resp.json()
                            if users:
                                assignee_id = users[0].get("id")
                except Exception as ae:
                    logger.warning("CSV import row %d: assignee lookup failed for '%s': %s", idx, assignee_username, ae)

            issue = gitlab_client.create_issue(
                title=title,
                description=description,
                labels=labels,
                project_id=pid,
                assignee_id=assignee_id,
            )
            imported += 1
        except Exception as e:
            logger.error("CSV import row %d failed: %s", idx, e)
            failed.append({"row": idx, "title": title, "error": str(e)})

    return {"imported": imported, "failed": failed}
